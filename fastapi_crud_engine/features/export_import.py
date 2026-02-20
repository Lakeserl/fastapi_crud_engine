from __future__ import annotations
import csv
import io
from typing import Any

from fastapi import UploadFile
from fastapi.responses import StreamingResponse


def export_csv(filename: str, items: list[Any]) -> StreamingResponse:
    if not items:
        buf = io.StringIO()
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
        )

    cols = _columns(items[0])

    def generate():
        buf    = io.StringIO()
        writer = csv.DictWriter(buf, fieldnames=cols)
        writer.writeheader()
        yield buf.getvalue()
        buf.seek(0)
        buf.truncate(0)
        for item in items:
            writer.writerow({c: _safe(_value(item, c)) for c in cols})
            yield buf.getvalue()
            buf.seek(0)
            buf.truncate(0)

    return StreamingResponse(
        generate(),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
    )


def export_excel(filename: str, items: list[Any]) -> StreamingResponse:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError(
            "openpyxl is required for Excel export. "
            "Install with: pip install 'fastapi-smart-crud[excel]'"
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename[:31]   

    if not items:
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return _xlsx_response(filename, buf)

    cols = _columns(items[0])
    ws.append(cols)
    for item in items:
        ws.append([_safe(_value(item, c)) for c in cols])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return _xlsx_response(filename, buf)


def _xlsx_response(filename: str, buf: io.BytesIO) -> StreamingResponse:
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


async def import_csv_or_excel(
    file: UploadFile,
    model_fields: list[str],
) -> tuple[list[dict], list[dict]]:
    content  = await file.read()
    filename = (file.filename or "").lower()

    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        rows = _parse_excel(content)
    else:
        rows = _parse_csv(content)

    valid:  list[dict] = []
    errors: list[dict] = []

    for i, row in enumerate(rows, start=2):  
        filtered = {k: v for k, v in row.items() if k in model_fields}
        if not filtered:
            errors.append({"row": i, "error": "No recognisable fields in row"})
            continue
        valid.append(filtered)

    return valid, errors


def _parse_csv(content: bytes) -> list[dict]:
    text   = content.decode("utf-8-sig")  
    reader = csv.DictReader(io.StringIO(text))
    return [dict(row) for row in reader]


def _parse_excel(content: bytes) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        raise RuntimeError("openpyxl required for Excel import.")

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows    = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]


def _safe(val: Any) -> Any:
    if val is None:
        return ""
    return str(val) if not isinstance(val, (int, float, bool)) else val


def _columns(item: Any) -> list[str]:
    if isinstance(item, dict):
        return [str(k) for k in item.keys()]
    mapper = getattr(item, "__mapper__", None)
    if mapper is None:
        raise RuntimeError("Export item must be ORM object or dict.")
    return [c.key for c in mapper.column_attrs]


def _value(item: Any, key: str) -> Any:
    if isinstance(item, dict):
        return item.get(key)
    return getattr(item, key, None)
